# from tools.calculating import calculate_data
from common.handle_log import do_log
from PIL import Image
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook
from matplotlib import rcParams
import threading
import pandas as pd
import matplotlib.pyplot as plt
import xlsxwriter
import subprocess
import re
import time
import csv
import pynvml
import datetime
import psutil
import matplotlib
import os


def plot_resource_usage(xlsx_path):
    """
    从文件中读取资源使用情况，绘图并保存图像，并将图像路径保存到xlsx中
    """
    # 读取数据
    matplotlib.use('Agg')
    try:
        df = pd.read_excel(xlsx_path)
    except Exception as e:
        do_log.error(f"[读取数据失败]：{e}")

    df['时间'] = pd.to_datetime(df['时间'], format='%H:%M:%S')

    # 设置中文字体
    plt.style.use('seaborn-v0_8-darkgrid')
    rcParams['font.sans-serif'] = ['Microsoft YaHei']
    rcParams['axes.unicode_minus'] = False

    fig, ax1 = plt.subplots(figsize=(12, 6))
    ax1.plot(df['时间'], df['GPU利用率/%'], label='GPU利用率/%', color='red')
    ax1.plot(df['时间'], df['CPU利用率/%'], label='CPU利用率/%', color='blue')
    ax1.set_ylabel('GPU / CPU 利用率（%）', color='black')

    ax2 = ax1.twinx()
    ax2.plot(df['时间'], df['显存使用量/Mib'], label='显存使用量/Mib', color='orange', linestyle='--')
    ax2.plot(df['时间'], df['内存使用量/MB'], label='内存使用量/MB', color='green', linestyle='--')
    ax2.plot(df['时间'], df['磁盘使用量/MB'], label='磁盘使用量/MB', color='purple', linestyle='--')
    ax2.set_ylabel('显存 / 内存 / 磁盘（MB）', color='black')

    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left')

    ax1.set_xlabel('Time')
    plt.title('资源使用情况（随时间变化）')
    plt.xticks(rotation=45)
    plt.tight_layout()

    # 保存图像
    image_path = os.path.splitext(xlsx_path)[0] + '_usage_plot.png'
    plt.savefig(image_path)
    plt.close()

    # 插入图像到 Excel
    try:
        if os.path.getsize(xlsx_path) < 1000:
            do_log.warning(f"{xlsx_path} 文件太小，可能数据采集失败")
            return
        wb = load_workbook(xlsx_path)
        ws = wb.active
        img = XLImage(image_path)
        ws.add_image(img, 'J6')  # 可以根据需要修改插入位置
        wb.save(xlsx_path)
        do_log.info(f"✅ 图像已插入 Excel 文件：{xlsx_path}")
    except Exception as e:
        do_log.error(f"插入图像失败：{e}")

def get_training_utilization(name, status, event, gpu_index=0):
    '''监控训练资源使用情况，并记录成 Excel'''
    do_log.debug(f"[{name}] 开始资源监控线程")
    start_time = time.time()
    file_name = name + '.xlsx'
    headers = ['时间', 'GPU利用率/%', '显存使用量/Mib', 'CPU利用率/%', '内存使用量/MB', '磁盘使用量/MB']

    try:
        workbook = xlsxwriter.Workbook(file_name)
        worksheet = workbook.add_worksheet()
        for i, h in enumerate(headers):
            worksheet.write(0, i, h)
    except Exception as e:
        do_log.error(f"创建Excel失败: {e}")
        event.set()
        return

    row = 1
    partition = "D:\\"  # 可根据实际改成项目盘符
    try:
        init_disk_used = psutil.disk_usage(partition).used / (1024 * 1024)
        do_log.info(f"[{name}] 初始磁盘使用量: {init_disk_used:.2f}MB")
    except Exception as e:
        do_log.warning(f"获取磁盘信息失败：{e}")
        init_disk_used = 0

    target_names = {'VDL.exe', 'python.exe', 'QtWebEngineProcess.exe'}
    target_pids = [proc.pid for proc in psutil.process_iter(['name']) if proc.info['name'] in target_names]
    for pid in target_pids:
        try:
            psutil.Process(pid).cpu_percent(None)
        except:
            continue

    try:
        while True:
            if status[0] == 1:
                do_log.debug(f"[{name}] 状态标志位为 1，退出资源监控")
                break

            # GPU采集
            try:
                output = subprocess.check_output([
                    "nvidia-smi",
                    "--query-gpu=utilization.gpu,memory.used",
                    f"--id={gpu_index}",
                    "--format=csv,noheader,nounits"
                ], timeout=2).decode()
                match = re.findall(r"(\d+),\s*(\d+)", output)
                gpu_util, mem_used = map(int, match[0]) if match else (0, 0)
            except Exception as e:
                do_log.warning(f"[{name}] GPU采集失败: {e}")
                gpu_util, mem_used = 0, 0

            # CPU和内存采集
            cpu_sum = mem_sum = 0.0
            for pid in target_pids:
                try:
                    proc = psutil.Process(pid)
                    cpu = proc.cpu_percent(interval=0.1) / psutil.cpu_count()
                    mem = proc.memory_info().rss / (1024 * 1024)
                    cpu_sum += cpu
                    mem_sum += mem
                except:
                    continue

            # 磁盘增量采集
            try:
                current_disk_used = psutil.disk_usage(partition).used / (1024 * 1024)
                delta_disk = current_disk_used - init_disk_used
            except:
                delta_disk = 0

            now_time = datetime.datetime.now().strftime('%H:%M:%S')

            # 写入Excel
            try:
                worksheet.write(row, 0, now_time)
                worksheet.write(row, 1, gpu_util)
                worksheet.write(row, 2, mem_used)
                worksheet.write(row, 3, round(cpu_sum, 2))
                worksheet.write(row, 4, round(mem_sum, 2))
                worksheet.write(row, 5, round(delta_disk, 2))
                row += 1
            except Exception as e:
                do_log.warning(f"[{name}] 写入Excel失败: {e}")
                break

            time.sleep(1)

    except Exception as e:
        do_log.error(f"[{name}] 资源监控主循环异常：{e}")
    finally:
        try:
            workbook.close()
            do_log.info(f"[{name}] 监控数据保存完成，共{row - 1}行")
        except Exception as e:
            do_log.error(f"[{name}] Excel关闭失败: {e}")

        # 通知调用方监控线程完成
        event.set()


# def get_infer_utilization(name,status,event,path):
#     '''获取推理阶段的资源使用情况'''
#     pynvml.nvmlInit()
#     current_dir = os.getcwd()

#     header = ['Time', 'GPU利用率/%', '显存使用量/Mib','CPU利用率/%','内存使用量/MB','磁盘使用量/GB']  #设置表头

#     file_name = name + '.csv'      
#     with open(file_name, 'a', newline='') as f:
#         writer = csv.writer(f)
#         writer.writerow(header)#将表头写入表格
#         '''磁盘使用情况'''              
#         partition = "D:\\"  #磁盘分区的挂载点路径
#         init_disk_used = psutil.disk_usage(partition).used / (1024 * 1024 * 1024)
#         print(f'磁盘已使用：{init_disk_used:.2f} GB')
#     if status[0] == 0:
#         while True:
#             #调用英伟达命令行
#             nvidia_smi_output = subprocess.check_output(["nvidia-smi","--query-gpu=utilization.gpu,memory.used", "--format=csv,noheader,nounits"]).decode("utf-8")
#             utilization = re.findall(r"(\d+),\s+(\d+)", nvidia_smi_output)

#             if len(utilization) > 0:
#                 gpu_util, mem_util = map(int, utilization[0]) #GPU利用率和显存使用量
                
#                 '''获取磁盘使用情况'''
#                 disk_used = psutil.disk_usage(partition).used / (1024 * 1024 * 1024)
#                 lated_disk_used = disk_used - init_disk_used
#                 print(f'磁盘使用量为：{lated_disk_used:.2f} GB')

#                 '''根据进程号获取CPU使用情况'''
#                 cpu_sum = 0
#                 men_cum = 0
#                 plist = ['VDL.exe','python.exe','QtWebEngineProcess.exe']
#                 #print('cpu核数：',psutil.cpu_count())

#                 for proc in psutil.process_iter(['name']):
#                     if proc.info['name'] in plist:
#                         pid = proc.pid
#                         print(proc.info['name'] ,"的进程号是：", pid)
#                         process_cpu_usage = proc.cpu_percent()/psutil.cpu_count()
#                         process_mem_usage = proc.memory_info().rss / (1024*1024)
#                         cpu_sum += process_cpu_usage
#                         men_cum += process_mem_usage
                        
#                 now = datetime.datetime.now()
#                 current_time = now.strftime("%H:%M:%S")#获取当前时间
               
#                 with open(file_name, 'a', newline='') as f:
#                     writer = csv.writer(f)
#                     writer.writerow([current_time, gpu_util, mem_util,round(cpu_sum,2),round(men_cum,2),round(lated_disk_used,2)])
#                     time.sleep(1)
#                 if status[0] == 1:                       
#                     break
#     event.set() 

#     '''统计表格数据'''
#     calculate_data(file_name)
#     static_path = os.path.join(path, 'static')
#     project_name = f"{name}"
#     detail_screenshot = os.path.join(static_path, f"{project_name}.png")
#     min_time = os.path.join(static_path, f"{project_name}_min.png")
#     max_time = os.path.join(static_path, f"{project_name}_max.png")
#     avg_time = os.path.join(static_path, f"{project_name}_avg.png")

#     airtest_method.screenshot(detail_screenshot)   #全屏截图
#     detail_image =  os.path.join(path,detail_screenshot)  #指定路径
#     region_image = Image.open(detail_image)

#     min_time_screenshot = region_image.crop((1440,428,1580,515)) #最小值截图
#     min_time_screenshot.save(min_time)

#     max_time_screenshot = region_image.crop((1600,428,1740,515)) #最大值截图
#     max_time_screenshot.save(max_time)

#     avg_time_screenshot = region_image.crop((1760,428,1900,515)) #平均值截图
#     avg_time_screenshot.save(avg_time)

#     ocr_character = [min_time,max_time,avg_time]
#     ocr_content = ocr_organize(ocr_character)    
#     # data = [int(ocr_content[0]),int(ocr_content[1])]
#     # van_time = statistics.variance(data)
#     # range_time = ocr_content[1] - ocr_content[0]      
        
#     '''将ct时间记录至表格'''
#     content = {}
#     content = {
#         'min:' :ocr_content[0],
#         'max:' :ocr_content[1],
#         'avg:' :ocr_content[2],
#         # 'van:' :van_time,
#         # 'range:' : range_time
#     }
#     with open(file_name, 'a', newline='') as f:
#         writer = csv.writer(f)
#         writer.writerow([content])
 
def monitor_memory_usage(threshold_percent=90, check_interval=1):
    """实时监控内存使用情况并进行预警
    Args:
        threshold_percent: 内存使用率预警阈值，默认90%
        check_interval: 检查间隔时间（秒），默认1秒
    """
    while True:
        try:
            # 获取内存使用情况
            memory = psutil.virtual_memory()
            memory_percent = memory.percent
            memory_used = memory.used / (1024 * 1024 * 1024)  # 转换为GB
            memory_total = memory.total / (1024 * 1024 * 1024)  # 转换为GB
            
            # 获取进程内存使用情况
            process_memory = 0
            plist = ['VDL.exe', 'python.exe', 'QtWebEngineProcess.exe']
            for proc in psutil.process_iter(['name', 'memory_info']):
                if proc.info['name'] in plist:
                    process_memory += proc.info['memory_info'].rss / (1024 * 1024 * 1024)  # 转换为GB
            
            # 记录当前内存使用情况
            current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            do_log.info(f"Memory Usage at {current_time}:")
            do_log.info(f"Total Memory: {memory_total:.2f} GB")
            do_log.info(f"Used Memory: {memory_used:.2f} GB")
            do_log.info(f"Memory Usage: {memory_percent}%")
            do_log.info(f"Process Memory: {process_memory:.2f} GB")
            
            # 检查内存使用率并发出不同级别的预警
            if memory_percent > 95:
                error_msg = f"Critical memory usage: {memory_percent}% (Threshold: 95%)"
                do_log.error(error_msg)
                # 如果进程内存使用过高，记录详细信息
                if process_memory > memory_total * 0.5:  # 进程使用超过总内存的50%
                    do_log.error(f"Critical process memory usage: {process_memory:.2f} GB")
                    for proc in psutil.process_iter(['name', 'memory_info']):
                        if proc.info['name'] in plist:
                            proc_mem = proc.info['memory_info'].rss / (1024 * 1024 * 1024)
                            do_log.error(f"Process {proc.info['name']} using {proc_mem:.2f} GB")
                # 抛出异常
                raise MemoryError(f"Memory usage exceeds critical threshold: {memory_percent}%")
                
            elif memory_percent > threshold_percent:
                warning_msg = f"Memory usage warning: {memory_percent}% (Threshold: {threshold_percent}%)"
                do_log.warning(warning_msg)
                # 如果进程内存使用过高，记录详细信息
                if process_memory > memory_total * 0.5:  # 进程使用超过总内存的50%
                    do_log.warning(f"High process memory usage: {process_memory:.2f} GB")
                    for proc in psutil.process_iter(['name', 'memory_info']):
                        if proc.info['name'] in plist:
                            proc_mem = proc.info['memory_info'].rss / (1024 * 1024 * 1024)
                            do_log.warning(f"Process {proc.info['name']} using {proc_mem:.2f} GB")
            
            time.sleep(check_interval)
            
        except MemoryError as e:
            do_log.error(f"Memory monitoring error: {str(e)}")
            raise  # 重新抛出异常
        except Exception as e:
            do_log.error(f"Error in memory monitoring: {str(e)}")
            time.sleep(check_interval)

def monitor_training_memory(threshold=90, interval=1):
    """监控训练过程中的内存使用情况
    Args:
        threshold: 内存使用率预警阈值，默认90%
        interval: 检查间隔时间（秒），默认1秒
    Returns:
        status: 控制监控状态的列表
    """
    status = [0]  # 用于控制监控状态
    
    def monitor():
        while status[0] == 0:
            try:
                # 获取内存使用情况
                memory = psutil.virtual_memory()
                memory_percent = memory.percent
                memory_used = memory.used / (1024 * 1024 * 1024)  # 转换为GB
                memory_total = memory.total / (1024 * 1024 * 1024)  # 转换为GB
                
                # 获取进程内存使用情况
                process_memory = 0
                plist = ['VDL.exe', 'python.exe', 'QtWebEngineProcess.exe']
                for proc in psutil.process_iter(['name', 'memory_info']):
                    if proc.info['name'] in plist:
                        process_memory += proc.info['memory_info'].rss / (1024 * 1024 * 1024)  # 转换为GB
                
                # 记录当前内存使用情况
                current_time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                do_log.info(f"Memory Usage at {current_time}:")
                do_log.info(f"Total Memory: {memory_total:.2f} GB")
                do_log.info(f"Used Memory: {memory_used:.2f} GB")
                do_log.info(f"Memory Usage: {memory_percent}%")
                do_log.info(f"Process Memory: {process_memory:.2f} GB")
                
                # 检查内存使用率并发出不同级别的预警
                if memory_percent > 95:
                    error_msg = f"Critical memory usage: {memory_percent}% (Threshold: 95%)"
                    do_log.error(error_msg)
                    # 如果进程内存使用过高，记录详细信息
                    if process_memory > memory_total * 0.5:  # 进程使用超过总内存的50%
                        do_log.error(f"Critical process memory usage: {process_memory:.2f} GB")
                        for proc in psutil.process_iter(['name', 'memory_info']):
                            if proc.info['name'] in plist:
                                proc_mem = proc.info['memory_info'].rss / (1024 * 1024 * 1024)
                                do_log.error(f"Process {proc.info['name']} using {proc_mem:.2f} GB")
                    # 抛出异常
                    raise MemoryError(f"Memory usage exceeds critical threshold: {memory_percent}%")
                    
                elif memory_percent > threshold:
                    warning_msg = f"Memory usage warning: {memory_percent}% (Threshold: {threshold}%)"
                    do_log.warning(warning_msg)
                    # 如果进程内存使用过高，记录详细信息
                    if process_memory > memory_total * 0.5:  # 进程使用超过总内存的50%
                        do_log.warning(f"High process memory usage: {process_memory:.2f} GB")
                        for proc in psutil.process_iter(['name', 'memory_info']):
                            if proc.info['name'] in plist:
                                proc_mem = proc.info['memory_info'].rss / (1024 * 1024 * 1024)
                                do_log.warning(f"Process {proc.info['name']} using {proc_mem:.2f} GB")
                
                time.sleep(interval)
                
            except MemoryError as e:
                do_log.error(f"Memory monitoring error: {str(e)}")
                raise
            except Exception as e:
                do_log.error(f"Error in memory monitoring: {str(e)}")
                time.sleep(interval)
    
    # 启动监控线程
    monitor_thread = threading.Thread(target=monitor, daemon=True)
    monitor_thread.start()
    
    return status

if __name__ == "__main__":
    plot_resource_usage(r"D:\ly\VDL_autotest\分类算法.csv")
        
    
